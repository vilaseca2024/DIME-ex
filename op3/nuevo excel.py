import re
import openpyxl
from PyPDF2 import PdfReader
from tkinter import filedialog
import os

# ------------------ PARTE 1: Extraer campos del PDF ------------------ #
campos = []

with open("DIM.pdf", "rb") as pdffileobj:
    pdfreader = PdfReader(pdffileobj)
    full_text = ""
    for page in pdfreader.pages:
        text = page.extract_text()
        if text:
            full_text += text + "\n"

# Detectar bloques A1. B2. H8.1 etc.
pattern = r'([A-Z](?:\d+(?:\.\d+)?))\.\s*([\s\S]*?)(?=(?:[A-Z](?:\d+(?:\.\d+)?))\.|$)'
matches = re.findall(pattern, full_text, re.DOTALL)

for id_actual, bloque in matches:
    bloque = bloque.strip()
    lines = bloque.split("\n")
    titulo = lines[0] if lines else ""
    valor = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""

    
    if id_actual in ["H8", "I2"]:
        subpattern = r'(\d{1,2})\s+(.+?)(?=(?:\d{1,2}\s+)|$)'
        submatches = re.findall(subpattern, valor)
        for num, subbloque in submatches:
            subbloque = subbloque.strip()
            sub_id = f"{id_actual}.{num}"
            campos.append({'id': sub_id, 'titulo': titulo.strip(), 'valor': subbloque})
    elif id_actual == "E1":
        # Buscar subcampos como "E1.1", "E1.2"
        subpattern = r'(E1\.\d+)\s+Título:\s*(.*?)\s*=>\s*Valor:\s*(.*?)(?=(?:E1\.\d+|E2\.)|$)'
        submatches = re.findall(subpattern, valor, re.DOTALL)
        # Extraer cuerpo principal de E1 (todo antes de la primera coincidencia)
        if submatches:
            idx = valor.find(submatches[0][0])
            valor_e1 = valor[:idx].strip()
            campos.append({'id': "E1", 'titulo': "Datos del Proveedor", 'valor': valor_e1})
            for sid, subtitulo, subvalor in submatches:
                campos.append({'id': sid.strip(), 'titulo': subtitulo.strip(), 'valor': subvalor.strip()})
        else:
            # Si no hay subcampos, agregar como bloque único
            campos.append({'id': id_actual, 'titulo': titulo.strip(), 'valor': valor.strip()})
    else:
        campos.append({'id': id_actual, 'titulo': titulo.strip(), 'valor': valor})


# Procesar sección "J." e "G."
if "J." in full_text:
    j_match = re.search(r'(J\.\s+Información adicional del Ítem.*?)H\.', full_text, re.DOTALL)
    if j_match:
        contenido = j_match.group(1)
        lines = contenido.strip().split("\n")
        oi_line = next((l for l in lines if l.startswith("OI")), None)
        if oi_line:
            campos.append({'id': 'J.', 'titulo': 'Información adicional del Ítem', 'valor': oi_line.strip()})

        for tipo in ['GA', 'IVA']:
            for l in lines:
                if l.startswith(tipo):
                    partes = l.split()
                    valor_final = partes[-1] if partes else ""
                    campos.append({'id': f'J.{tipo}', 'titulo': tipo, 'valor': valor_final})

if "G." in full_text:
    g_match = re.search(r'(G\.\s+Observaciones generales.*?)Total tributos a pagar\s+(\d+)', full_text, re.DOTALL)
    if g_match:
        contenido, total = g_match.groups()
        lines = contenido.strip().split("\n")
        obs_line = lines[1].strip() if len(lines) > 1 else ""
        campos.append({'id': 'G.', 'titulo': 'Observaciones generales de la declaración', 'valor': obs_line})
        for tipo in ['GA', 'IVA']:
            for l in lines:
                if l.startswith(tipo):
                    partes = l.split()
                    valor_final = partes[-1] if partes else ""
                    campos.append({'id': f'G.{tipo}', 'titulo': tipo, 'valor': valor_final})
        campos.append({'id': 'G.TOTAL', 'titulo': 'Total tributos a pagar', 'valor': total.strip()})

# ------------------ PARTE 2: Procesar el Excel ------------------ #
 

with open("campos89.txt", "w", encoding="utf-8") as out:
    for c in campos:
        out.write(f"{c['id']}  Título: {c['titulo']}  =>  Valor: {c['valor']}\n")

print(f"Exportados {len(campos)} campos a 'campos.txt'")


ruta_excel = filedialog.askopenfilename(title="Selecciona el archivo Excel", filetypes=[("Excel files", "*.xlsx")])
if not ruta_excel or not os.path.exists(ruta_excel):
    print("No se seleccionó archivo válido.")
    exit()

wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

# Crear diccionario de búsqueda {id: valor}
dict_campos = {c['id']: c['valor'] for c in campos}

# Leer primera fila
for col in ws.iter_cols(min_row=1, max_row=1):
    celda = col[0]
    contenido = celda.value
    if not contenido:
        continue

    # Caso fórmula (ej. A2 + H8.3)
    if '+' in contenido:
        partes = [p.strip() for p in contenido.split('+')]
        suma = 0
        for p in partes:
            val = dict_campos.get(p)
            try:
                suma += float(val)
            except:
                pass
        ws.cell(row=3, column=celda.column, value=suma)

    # Valor directo
    else:
        valor = dict_campos.get(contenido)
        if valor is not None:
            # Si es número, lo escribimos como número
            try:
                valor = float(valor)
            except:
                pass
            ws.cell(row=3, column=celda.column, value=valor)

# Guardar sin modificar formato
nombre_guardado = os.path.splitext(os.path.basename(ruta_excel))[0] + "_completado.xlsx"
wb.save(nombre_guardado)

print(f"Excel procesado y guardado como '{nombre_guardado}'")
