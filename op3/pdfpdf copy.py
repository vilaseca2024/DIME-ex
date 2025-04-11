import re
from PyPDF2 import PdfReader

pdffileobj = open("DIM.pdf", "rb")
pdfreader = PdfReader(pdffileobj)
full_text = ""
for page in pdfreader.pages:
    full_text += page.extract_text() + "\n"
print(full_text)

def cargar():
    import json
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # Función para llenar el Excel con los datos extraídos del JSON
    def procesar_json_a_excel(json_data, excel_file):
        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Procesados"

        # Encabezados para las columnas
        encabezados = [
            "Descripción 2", "Unidad De Medida", "Cantidad Pedida", "Costo Unitario", 
            "Costo Total", "Moneda", "Fecha Entrega Prometida", "Descripción Criticidad",
            "Tipo de Línea", "NUMERO DE PARCIAL", "Cantidad Pedida", "Cantidad Recibida", 
            "Cantidad Pendiente", "Flete", "Otros Gastos"
        ]
        
        # Escribir los encabezados en la primera fila
        for col_num, header in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
        
        # Extraer datos del JSON
        try:
            datos_mercancia = json_data["datosMercancias"][0]
            descripcion = datos_mercancia["descripcionMercanciaComercial"]["ForPre"]  # "2880000 LITROS A GRANELEN CISTERNAS"
            cantidad_pedida = int(descripcion.split()[0])  # Extraemos 2880000
            precio_unitario = datos_mercancia["precioUnitario"]
            moneda = json_data["datosTransacciones"][0]["monedaTransaccion"]["descripcion"]
            fecha_embarque = json_data["lugares"]["fechaEmbarque"]
            
            # Cálculo de costo total
            costo_total = cantidad_pedida * precio_unitario
            
            # Escribir los datos en la fila 2
            fila = [
                descripcion,  # Descripción 2
                datos_mercancia["unidadMedida"],  # Unidad De Medida
                cantidad_pedida,  # Cantidad Pedida
                precio_unitario,  # Costo Unitario
                costo_total,  # Costo Total
                moneda,  # Moneda
                fecha_embarque,  # Fecha Entrega Prometida
                "Regular",  # Descripción Criticidad (ejemplo)
                "S",  # Tipo de Línea (ejemplo)
                "1",  # NUMERO DE PARCIAL (ejemplo, se debe modificar según la lógica)
                cantidad_pedida,  # Cantidad Pedida (nuevamente, depende de la lógica)
                cantidad_pedida,  # Cantidad Recibida (esto debe depender de los datos)
                cantidad_pedida,  # Cantidad Pendiente (esto también depende de los datos)
                0,  # Flete (puedes modificar si tienes los datos)
                0  # Otros Gastos (puedes modificar si tienes los datos)
            ]
            
            # Escribir los datos en la segunda fila
            for col_num, value in enumerate(fila, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}2'] = value
            
            # Guardar el archivo Excel
            wb.save(excel_file)
            print(f"Archivo Excel generado exitosamente: {excel_file}")
        
        except KeyError as e:
            print(f"Error al procesar el JSON: la clave {e} no se encuentra en el archivo.")

    # Función principal para leer el JSON y procesarlo
    def main():
        # Lee el archivo JSON (puedes modificar la ruta aquí)
        json_file = 'datos.json'  # Cambia el nombre del archivo JSON si es necesario
        excel_file = 'datos_procesados.xlsx'  # Nombre del archivo Excel de salida

        # Cargar el JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Llamar a la función que procesa el JSON y llena el Excel
        procesar_json_a_excel(json_data, excel_file)
    