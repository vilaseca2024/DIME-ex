import re
import openpyxl
from PyPDF2 import PdfReader
import os
from tkinter import Tk, Label, Button, filedialog, messagebox

sublevel_ids = {'H8.', 'E1.', 'I2.'}

class PDFProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de DIM a formato Excel")
        self.campos = []
        self.ruta_excel = None

        self.label_estado = Label(root, text="Cargue el archivo PDF para comenzar.")
        self.label_estado.pack(pady=10)

        self.boton_cargar_pdf = Button(root, text="Cargar PDF", command=self.cargar_pdf)
        self.boton_cargar_pdf.pack(pady=5)

        self.label_excel = Label(root, text="")
        self.label_excel.pack(pady=10)

        self.boton_procesar = Button(root, text="Procesar", command=self.procesar, state="disabled")
        self.boton_procesar.pack(pady=5)

    def cargar_pdf(self):
        ruta_pdf = filedialog.askopenfilename(title="Selecciona el archivo PDF", filetypes=[("PDF files", "*.pdf")])
        if not ruta_pdf:
            return

        
        self.campos = self.extraer_campos_pdf(ruta_pdf)
        self.label_estado.config(text="PDF cargado y campos extraídos.")

        self.ruta_excel = filedialog.askopenfilename(title="Selecciona el archivo Excel (mosol)", filetypes=[("Excel files", "*.xlsx")])
        if not self.ruta_excel:
            messagebox.showwarning("Error", "No se seleccionó el archivo Excel.")
            return

        self.label_excel.config(text="Archivo mosol para generar informe cargado correctamente.")
        self.boton_procesar.config(state="normal")

    def extraer_campos_pdf(self, ruta_pdf):
        campos = []
        with open(ruta_pdf, "rb") as pdffileobj:
            pdfreader = PdfReader(pdffileobj)
            full_text = ""
            for page in pdfreader.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

        pattern = r'([A-Z]\d*(?:\.\d+)?\.)\s*([\s\S]*?)(?=(?:[A-Z]\d*(?:\.\d+)?\.)|$)'
        matches = re.findall(pattern, full_text, re.DOTALL)

        for id_actual, bloque in matches:
            lines = [l.strip() for l in bloque.splitlines() if l.strip()]
            if not lines:
                continue

            if id_actual in {'J.', 'G.'}:
                campos.append({'id': id_actual, 'titulo': lines[0], 'valor': lines[1] if len(lines) > 1 else ""})
                for line in lines[2:]:
                    if re.match(r'^(Liquidación|Tipo|Sub totales)', line): continue
                    m = re.match(r'^(GA|IVA)\b', line)
                    if m:
                        code = m.group(1)
                        last_val = line.split()[-1]
                        campos.append({'id': f"{id_actual.rstrip('.')}.{code}", 'titulo': code, 'valor': last_val})
                        continue
                    if id_actual == 'G.' and line.startswith("Total tributos a pagar"):
                        parts = line.split()
                        campos.append({'id': f"{id_actual.rstrip('.')}.TOTAL", 'titulo': " ".join(parts[:-1]), 'valor': parts[-1]})
                continue

            if id_actual in sublevel_ids:
                base = id_actual.rstrip('.')
                current_id = ""
                current_title = ""
                current_value = ""
                last_num = -1

                for line in lines:
                    m = re.match(r'^(\d+)\s+(.*)', line)
                    if m:
                        num = int(m.group(1))
                        title_or_value = m.group(2).strip()

                        if num > last_num:
                            if current_id:
                                campos.append({'id': current_id, 'titulo': current_title, 'valor': current_value.strip()})
                            current_id = f"{base}.{num}"
                            current_title = title_or_value
                            current_value = ""
                            last_num = num
                        else:
                            current_value += " " + line.strip()
                    else:
                        current_value += " " + line.strip()

                if current_id:
                    campos.append({'id': current_id, 'titulo': current_title, 'valor': current_value.strip()})
                continue

            titulo = lines[0]
            valor = " ".join(lines[1:]).strip() if len(lines) > 1 else ""
            campos.append({'id': id_actual, 'titulo': titulo, 'valor': valor})

        for c in campos:
            if c['valor']:
                m = re.match(r'^([^\d,]*\([\w\s]+\))\s+([\d,\.]+)$', c['valor'])
                if m:
                    c['titulo'] += " " + m.group(1).strip()
                    c['valor'] = m.group(2).strip()

        clean_campos = []
        prev = None
        for c in campos:
            if c['id'] == 'A.' and prev and prev['id'] in {'B1.', 'B2.'}:
                if prev['valor']:
                    prev['valor'] += " " + c['titulo'] + " " + c['valor']
                else:
                    prev['valor'] = c['titulo'] + " " + c['valor']
            else:
                clean_campos.append(c)
                prev = c
        campos = clean_campos

        for c in campos:
            if c['id'] == 'H8.9':
                if not c['valor']:
                    idx = campos.index(c)
                    if idx + 1 < len(campos):
                        next_c = campos[idx + 1]
                        m = re.match(r'^H8\.(\d+)$', next_c['id'])
                        if m:
                            c['valor'] = m.group(1)
                else:
                    m = re.match(r'(\d+)', c['valor'])
                    if m:
                        c['valor'] = m.group(1)
       

        with open("campsoAA.txt", "w", encoding="utf-8") as out:    
            for c in campos:
                out.write(f"{c['id']}  Título: {c['titulo']}  =>  Valor: {c['valor']}\n")

        print(f"Exportados {len(campos)} campos a 'campsoAA.txt'")

        return campos


    def procesar(self):
        campos = self.campos
        dict_campos = {c['id']: c['valor'] for c in campos}

        try:
            wb = openpyxl.load_workbook(self.ruta_excel)
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un error no se pudo abrir el Excel: {str(e)}")
            return

        ws = wb.active

        for col in ws.iter_cols(min_row=1, max_row=1):
            celda = col[0]
            contenido = celda.value
            if not contenido:
                continue

            if '+' in str(contenido):
                partes = [p.strip() for p in contenido.split('+')]
                suma = 0
                for p in partes:
                    val = dict_campos.get(p)
                    try:
                        suma += float(val)
                    except:
                        pass
                row = 3
                while ws.cell(row=row, column=celda.column).value is not None:
                    row += 1
                ws.cell(row=row, column=celda.column, value=suma)
            else:
                valores = [c['valor'] for c in campos if c['id'] == contenido]
                row = 3
                for val in valores:
                    try:
                        val = float(val)
                    except:
                        pass
                    while ws.cell(row=row, column=celda.column).value is not None:
                        row += 1
                    ws.cell(row=row, column=celda.column, value=val)

        ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if ruta_guardado:
            wb.save(ruta_guardado)
            messagebox.showinfo("Éxito", f"Excel procesado y guardado en:\n{ruta_guardado}")
        else:
            messagebox.showwarning("Cancelado", "No se guardó el archivo.")




if __name__ == "__main__":
    root = Tk()
    root.geometry('520x300')
    root.resizable(0,1)
    app = PDFProcessorApp(root)
    root.mainloop()
    
