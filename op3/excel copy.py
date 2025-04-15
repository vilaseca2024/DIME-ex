import re
import openpyxl
from PyPDF2 import PdfReader
import json
import pandas as pd
from tkinter import Tk, Button, Label, filedialog, messagebox
import os
from datetime import datetime



pdf_path = 'DIM2.pdf'
template_path = 'mosol.xlsx'
output_path = 'mosol_output_fixed7.xlsx'
reader = PdfReader(open(pdf_path, 'rb'))
full_text = ""
for page in reader.pages:
    text = page.extract_text()
    if text:
        full_text += text + "\n"

pattern = r'([A-Z]\d*(?:\.\d+)*)\.\s*([\s\S]*?)(?=(?:[A-Z]\d*(?:\.\d+)*)\.|$)'
matches = re.findall(pattern, full_text, re.DOTALL)
raw_mapping = {}
for id_actual, bloque in matches:
    id_actual = id_actual.strip() + '.'
    bloque = bloque.strip()
    raw_mapping[id_actual] = bloque

parsed_mapping = {}
for id_actual, bloque in raw_mapping.items():
    if id_actual == 'B1.':
        after_colon = bloque.split(':', 1)[1].strip() if ':' in bloque else bloque
        parts = after_colon.split()
        parsed_mapping['B1._1'] = parts[0] if len(parts) >= 1 else ''
        parsed_mapping['B1._2'] = parts[1] if len(parts) >= 2 else ''
    elif id_actual == 'E1.':
        after_colon = bloque.split(':', 1)[1].strip() if ':' in bloque else bloque
        supplier_name = after_colon.split(',', 1)[0].strip()
        parsed_mapping['E1.'] = supplier_name
    else:
        if '\n' in bloque:
            _, valor = bloque.split('\n', 1)
            parsed_mapping[id_actual] = valor.strip()
        else:
            if ':' in bloque:
                parsed_mapping[id_actual] = bloque.split(':', 1)[1].strip()
            else:
                parsed_mapping[id_actual] = bloque

wb = openpyxl.load_workbook(template_path)
ws = wb.active

id_counts = {}
for col in range(1, ws.max_column + 1):
    cell_id = ws.cell(row=1, column=col).value
    if not cell_id:
        continue
    id_key = str(cell_id).strip()
    if not id_key.endswith('.'):
        id_key = id_key + '.'
    id_counts[id_key] = id_counts.get(id_key, 0) + 1
    count = id_counts[id_key]
    if id_key == 'B1.':
        if count == 1:
            value = parsed_mapping.get('B1._1', '')
        elif count == 2:
            value = parsed_mapping.get('B1._2', '')
        else:
            value = ''
    else:
        value = parsed_mapping.get(id_key, '')
    
    ws.cell(row=3, column=col, value=value)

wb.save(output_path)

output_path


def cargar_json():
    ruta = filedialog.askopenfilename(title="Selecciona un archivo JSON", filetypes=[("Archivos JSON", "*.json")])
    if ruta:
        with open(ruta, 'r', encoding='utf-8-sig') as f:
            datos = json.load(f)
 
       
# Interfaz
ventana = Tk()
ventana.title("Procesador de PDF a Excel")
ventana.geometry("400x400")

Label(ventana, text="Conversor DIM ➜ Excel (VLSC)", font=("Arial", 14)).pack(pady=20)
Button(ventana, text="Cargar y Procesar PDF", command=cargar_json, width=30, height=2).pack()

ventana.mainloop()