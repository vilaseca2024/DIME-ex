import re
import os
import fitz  # PyMuPDF
import openpyxl
from PyPDF2 import PdfReader
from tkinter import Tk, Label, Button, filedialog, messagebox
from collections import defaultdict

sublevel_ids = {'H8.', 'E1.', 'I2.'}

class PDFProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de DIM a formato Excel")
        self.campos = []
        self.ruta_excel = None

        self.label_estado = Label(root, text="Cargue el archivo PDF para comenzar.")
        self.label_estado.pack(pady=10)

        self.boton_cargar_pdf = Button(root, text="Cargar PDF", command=self.cargar_pdf)
        self.boton_cargar_pdf.pack(pady=5)

        self.label_excel = Label(root, text="")
        self.label_excel.pack(pady=10)

        self.boton_procesar = Button(root, text="Procesar", command=self.procesar, state="disabled")
        self.boton_procesar.pack(pady=5)

    def cargar_pdf(self):
        ruta_pdf = filedialog.askopenfilename(
            title="Selecciona el archivo PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if not ruta_pdf:
            return

        try:
            self.campos = self.extraer_campos_pdf(ruta_pdf)
            self.label_estado.config(text="PDF procesado y campos extraídos.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el PDF: {e}")
            return

        self.ruta_excel = filedialog.askopenfilename(
            title="Selecciona el archivo Excel (mosol)",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not self.ruta_excel:
            messagebox.showwarning("Error", "No se seleccionó el archivo Excel.")
            return

        self.label_excel.config(text="Archivo mosol para generar informe cargado correctamente.")
        self.boton_procesar.config(state="normal")

    def extraer_campos_pdf(self, ruta_pdf):
        # 1) Redactar encabezado e insertar texto en cada página
        campos=[]
        doc = fitz.open(ruta_pdf)
        zona_encabezado = fitz.Rect(0, 0, doc[0].rect.width, 60)
        for pagina in doc:
            pagina.add_redact_annot(zona_encabezado, fill=(1, 1, 1))
            pagina.apply_redactions()
            pagina.insert_text(
                point=(zona_encabezado.x0 + 10, zona_encabezado.y0 + 15),
                text="Nueva pagina ojo",
                fontsize=10,
                fontname="helv",
                fill=(0, 0, 0),
            )
        ruta_redacted = ruta_pdf.replace('.pdf', '_redacted.pdf')
        doc.save(ruta_redacted)
        doc.close()

        # 2) Leer texto del PDF ya procesado
        full_text = ""
        with open(ruta_redacted, 'rb') as f:
            reader = PdfReader(f)
            for page in reader.pages:
                txt = page.extract_text()
                if txt:
                    full_text += txt + "\n"

        # 3) Extraer bloques por ID
        pattern = r'([A-Z]\d*(?:\.\d+)?\.)\s*([\s\S]*?)(?=(?:[A-Z]\d*(?:\.\d+)?\.)|$)'
        matches = re.findall(pattern, full_text, re.DOTALL)

        for id_actual, bloque in matches:
            lines = [l.strip() for l in bloque.splitlines() if l.strip()]
            if not lines:
                continue

            if id_actual in {'J.', 'G.'}:
                campos.append({'id': id_actual, 'titulo': lines[0], 'valor': lines[1] if len(lines) > 1 else ""})
                for line in lines[2:]:
                    if re.match(r'^(Liquidación|Tipo|Sub totales)', line): continue
                    m = re.match(r'^(GA|IVA)\b', line)
                    if m:
                        code = m.group(1)
                        last_val = line.split()[-1]
                        campos.append({'id': f"{id_actual.rstrip('.')}.{code}", 'titulo': code, 'valor': last_val})
                        continue
                    if id_actual == 'G.' and line.startswith("Total tributos a pagar"):
                        parts = line.split()
                        campos.append({'id': f"{id_actual.rstrip('.')}.TOTAL", 'titulo': " ".join(parts[:-1]), 'valor': parts[-1]})
                continue

            # Subniveles E1., H8., I2.
            if id_actual in sublevel_ids:
                base = id_actual.rstrip('.')
                base_lines, rest = [], []
                found_num=False
                for line in lines:
                    if re.match(r'^\d+\s+', line): found_num=True
                    if not found_num:
                        base_lines.append(line)
                    else:
                        rest.append(line)
                if base_lines:
                    campos.append({'id': id_actual, 'titulo': base_lines[0], 'valor': ' '.join(base_lines[1:]).strip() if len(base_lines)>1 else ''})
                cur_id=cur_title=cur_val=''
                last_num=-1
                for line in rest:
                    m=re.match(r'^(\d+)\s+(.*)', line)
                    if m:
                        num=int(m.group(1)); txt=m.group(2).strip()
                        if num>last_num:
                            if cur_id:
                                campos.append({'id':cur_id,'titulo':cur_title,'valor':cur_val.strip()})
                            cur_id=f"{base}.{num}"; cur_title=txt; cur_val=''; last_num=num
                        else:
                            cur_val += ' '+line.strip()
                    else:
                        cur_val += ' '+line.strip()
                if cur_id:
                    campos.append({'id':cur_id,'titulo':cur_title,'valor':cur_val.strip()})
                continue

            # resto campos
            titulo = lines[0]
            valor = ' '.join(lines[1:]).strip() if len(lines)>1 else ''
            campos.append({'id':id_actual,'titulo':titulo,'valor':valor})

        # 4) Mover (USD)
        for c in campos:
            if c['valor']:
                m=re.match(r'^([^\d,]*\([\w\s]+\))\s+([\d,\.]+)$',c['valor'])
                if m:
                    c['titulo'] += ' '+m.group(1).strip()
                    c['valor'] = m.group(2).strip()

        # 5) A. en B1./B2.
        clean=[]; prev=None
        for c in campos:
            if c['id']=='A.' and prev and prev['id'] in {'B1.','B2.'}:
                prev['valor'] = (prev['valor']+' '+c['titulo']+' '+c['valor']).strip()
            else:
                clean.append(c); prev=c
        campos=clean

        for c in campos:
            if c['id'] == 'H8.9':
                if not c['valor']:
                    idx = campos.index(c)
                    if idx + 1 < len(campos):
                        next_c = campos[idx + 1]
                        m = re.match(r'^H8\.(\d+)$', next_c['id'])
                        if m:
                            c['valor'] = m.group(1)
                else:
                    m = re.match(r'(\d+)', c['valor'])
                    if m:
                        c['valor'] = m.group(1)

        # 7) Distribuir valores si 'Nueva pagina ojo'
        for idx,c in enumerate(campos):
            if 'Nueva pagina ojo' in c['titulo']:
                if idx>=4 and all(campos[idx-k-1]['valor']=='' for k in range(4)):
                    vals=c['valor'].split()
                    if len(vals)==5 and all(re.match(r'^\d+(?:\.\d+)?$',v) for v in vals):
                        for j in range(4): campos[idx-4+j]['valor']=vals[j]
                        campos[idx]['valor']=vals[-1]
                    else:
                        nums=re.findall(r'\d+(?:\.\d+)?',c['valor'])
                        if nums:
                            first=re.search(r'\d+(?:\.\d+)?',c['valor']); last=None
                            for m in re.finditer(r'\d+(?:\.\d+)?',c['valor']): last=m
                            prefix=c['valor'][:first.start()].strip(); suffix=c['valor'][last.end():].strip()
                            segs=[prefix]+nums+[suffix]
                            if len(segs)==4:
                                for j in range(4): campos[idx-4+j]['valor']=segs[j]
                                campos[idx]['valor']=''

        # eliminar PDF temporal
        try: os.remove(ruta_redacted)
        except: pass
        with open("camposN.txt", "w", encoding="utf-8") as out:
            for c in campos:
                out.write(f"{c['id']}  Título: {c['titulo']}  =>  Valor: {c['valor']}\n")

        return campos

    def procesar(self):
        # Agrupar valores por ID
        dict_campos = defaultdict(list)
        for c in self.campos:
            dict_campos[c['id']].append(c['valor'])

        try:
            wb = openpyxl.load_workbook(self.ruta_excel)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el Excel: {e}")
            return
        ws = wb.active

        for col in ws.iter_cols(min_row=1, max_row=1):
            header = col[0].value
            if not header:
                continue

            if '+' in str(header):
                parts = [p.strip() for p in str(header).split('+')]
                # sumar todos los valores de cada parte
                total = 0.0
                for p in parts:
                    for val in dict_campos.get(p, []):
                        try:
                            total += float(val)
                        except:
                            pass
                row = 3
                while ws.cell(row=row, column=col[0].column).value is not None:
                    row += 1
                ws.cell(row=row, column=col[0].column, value=total)
            else:
                # escribir cada valor en fila nueva
                vals = dict_campos.get(str(header), [])
                row = 3
                for v in vals:
                    try:
                        v_float = float(v)
                    except:
                        v_float = v
                    while ws.cell(row=row, column=col[0].column).value is not None:
                        row += 1
                    ws.cell(row=row, column=col[0].column, value=v_float)

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if save_path:
            wb.save(save_path)
            messagebox.showinfo("Éxito", f"Guardado en: {save_path}")
        else:
            messagebox.showwarning("Cancelado", "No se guardó el archivo.")

if __name__ == '__main__':
    root = Tk()
    root.geometry('520x300')
    root.resizable(0,1)
    app = PDFProcessorApp(root)
    root.mainloop()
